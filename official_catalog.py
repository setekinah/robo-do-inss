"""Catálogo regulatório versionado a partir de documentos oficiais do Portal IN.

Este módulo nunca ativa regras por conta própria. Ele apenas coleta fontes da
lista permitida, preserva evidências e prepara versões para revisão jurídica.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from runtime_paths import DATA_DIR


MAX_SOURCE_BYTES = 25 * 1024 * 1024
OFFICIAL_SOURCE_REGISTRY = (
    {
        "key": "pt990_anexo_v",
        "title": "PT 990 — Anexo V: Relação dos Indicadores Disponibilizados no CNIS",
        "scope": "Indicadores CNIS",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt990/AnexoV.pdf",
    },
    {
        "key": "pt990_anexo_iii",
        "title": "PT 990 — Anexo III: Ajustes de complementação, utilização e agrupamento",
        "scope": "Ajustes EC 103/2019",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt990/AnexoIII.pdf",
    },
    {
        "key": "in128_rac",
        "title": "IN 128 — Anexo I: RAC completo",
        "scope": "Acertos de CNIS",
        "url": "https://portalin.inss.gov.br/assets/anexos/in/AnexoI.docx",
    },
    {
        "key": "in128_rac_vinculos",
        "title": "IN 128 — RAC 2.2: Acerto de vínculos e remunerações",
        "scope": "CNIS — vínculos",
        "url": "https://portalin.inss.gov.br/assets/anexos/in/AnexoI-B-2.2.docx",
    },
    {
        "key": "in128_rac_contribuicoes",
        "title": "IN 128 — RAC 2.6: Acerto de contribuições",
        "scope": "CNIS — contribuições",
        "url": "https://portalin.inss.gov.br/assets/anexos/in/AnexoI-F-2.6.docx",
    },
    {
        "key": "pt991_idade_mulher",
        "title": "PT 991 — Transição por idade da mulher",
        "scope": "Aposentadoria",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt991/AnexoIV.docx",
    },
    {
        "key": "pt991_art15",
        "title": "PT 991 — Transição por tempo, art. 15",
        "scope": "Aposentadoria",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt991/AnexoV.docx",
    },
    {
        "key": "pt991_art16",
        "title": "PT 991 — Transição por tempo, art. 16",
        "scope": "Aposentadoria",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt991/AnexoVII.docx",
    },
    {
        "key": "pt991_pedagio50",
        "title": "PT 991 — Transição com pedágio de 50%",
        "scope": "Aposentadoria",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt991/AnexoIX.docx",
    },
    {
        "key": "pt991_pedagio100",
        "title": "PT 991 — Transição com pedágio de 100%",
        "scope": "Aposentadoria",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt991/AnexoX.docx",
    },
    {
        "key": "in128_enquadramento_especial",
        "title": "IN 128 — Enquadramento de atividade especial",
        "scope": "Especial",
        "url": "https://portalin.inss.gov.br/assets/anexos/in/AnexoXVI.docx",
    },
    {
        "key": "in128_ppp",
        "title": "IN 128 — Perfil Profissiográfico Previdenciário",
        "scope": "Especial",
        "url": "https://portalin.inss.gov.br/assets/anexos/in/AnexoXVII.docx",
    },
    {
        "key": "in128_pcd_conversao",
        "title": "IN 128 — Tabela de conversão LC 142/2013",
        "scope": "PcD",
        "url": "https://portalin.inss.gov.br/assets/anexos/in/AnexoXVIII.docx",
    },
    {
        "key": "pt992_dependentes",
        "title": "PT 992 — Pensão por morte e auxílio-reclusão",
        "scope": "Dependentes",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt992/AnexoV.docx",
    },
    {
        "key": "pt992_cotas_dependentes",
        "title": "PT 992 — Duração das cotas de dependentes",
        "scope": "Pensão por morte",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt992/AnexoIII.docx",
    },
    {
        "key": "pt1208_social",
        "title": "PT 1.208 — Avaliação social",
        "scope": "BPC/LOAS",
        "url": "https://portalin.inss.gov.br/assets/anexos/pt1208/AnexoII.doc",
    },
)


@dataclass(frozen=True)
class ImportedCatalog:
    source_name: str
    source_url: str
    source_hash: str
    definitions: list[dict[str, Any]]


def _normalise_header(value: Any) -> str:
    value = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in value if not unicodedata.combining(char)).casefold().strip()


def _safe_cell(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def import_indicator_workbook(file_path: Path) -> ImportedCatalog:
    """Lê somente uma planilha local enviada pelo escritório para revisão."""
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - depende do ambiente de execução
        raise RuntimeError("Importação XLSX requer a dependência openpyxl.") from error

    content = file_path.read_bytes()
    if not content or len(content) > MAX_SOURCE_BYTES:
        raise ValueError("Planilha vazia ou maior que 25 MB.")
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Indicadores CNIS"] if "Indicadores CNIS" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("A planilha não possui cabeçalho.")
        positions = {_normalise_header(name): index for index, name in enumerate(headers) if name}
        required = {"tipo", "grupo", "sigla", "descricao oficial"}
        if not required.issubset(positions):
            raise ValueError("Planilha incompatível: esperadas as colunas Tipo, Grupo, Sigla e Descrição oficial.")

        definitions: list[dict[str, Any]] = []
        for row in rows:
            code = _safe_cell(row[positions["sigla"]] if positions["sigla"] < len(row) else "")
            description = _safe_cell(row[positions["descricao oficial"]] if positions["descricao oficial"] < len(row) else "")
            if not code or not description:
                continue
            source_url = _safe_cell(row[positions.get("fonte oficial", -1)] if positions.get("fonte oficial", -1) >= 0 else "")
            page = _safe_cell(row[positions.get("pagina do pdf", -1)] if positions.get("pagina do pdf", -1) >= 0 else "")
            definitions.append({
                "code": code.upper(),
                "indicator_type": _safe_cell(row[positions["tipo"]]),
                "indicator_group": _safe_cell(row[positions["grupo"]]),
                "official_description": description,
                "official_clarification": _safe_cell(row[positions.get("esclarecimentos oficiais", -1)] if positions.get("esclarecimentos oficiais", -1) >= 0 else ""),
                "general_guidance": _safe_cell(row[positions.get("orientacao geral", -1)] if positions.get("orientacao geral", -1) >= 0 else ""),
                "source_url": source_url or OFFICIAL_SOURCE_REGISTRY[0]["url"],
                "source_page": page,
                "canonical_key": hashlib.sha256(f"{code}|{description}".encode("utf-8")).hexdigest()[:24],
            })
    finally:
        workbook.close()
    if not definitions:
        raise ValueError("Nenhum indicador válido foi encontrado na planilha.")
    return ImportedCatalog(
        source_name=file_path.name,
        source_url=OFFICIAL_SOURCE_REGISTRY[0]["url"],
        source_hash=hashlib.sha256(content).hexdigest(),
        definitions=definitions,
    )


def _indicator_rows_from_pdf_pages(page_texts: list[str], source_url: str) -> list[dict[str, Any]]:
    """Extract auditable rows from the official PT 990 Annex V text layer.

    The PDF is a visual table, but its native text stream is column-linear.
    Rows are accepted only when a ``CsPendencia``, ``CsAlerta`` or ``CsAcerto``
    marker immediately precedes an indicator code. This prevents headings and
    narrative acronyms from entering the catalog as rules.
    """
    marker_pattern = re.compile(r"\b(CsPendencia|CsAlerta|CsAcerto)\b", re.IGNORECASE)
    code_pattern = re.compile(r"(?m)^\s*([A-Z][A-Z0-9]{2,}(?:-[A-Z0-9]+){0,5})\s*$")
    excluded_codes = {
        "TIPO", "GRUPO", "SIGLA", "DESCRICAO", "ESCLARECIMENTOS", "CNIS", "INSS",
        "CPF", "CNPJ", "NIT", "PIS", "SEI", "GPS", "DARF", "EC", "RPS",
    }
    definitions: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    explanation_starts = re.compile(
        r"^(?:O indicador|A pend[êe]ncia|Trata-se|Indicador|Este indicador|"
        r"O registro|Refere-se|Sinaliza|Informa|Aponta)\b",
        re.IGNORECASE,
    )

    for page_number, raw_page in enumerate(page_texts, start=1):
        text = raw_page or ""
        # O PDF quebra algumas siglas longas exatamente após o hífen
        # (por exemplo, ``PREC-\nFACULTCONC``). Reconstruímos somente linhas
        # integralmente em maiúsculas, antes de procurar códigos.
        text = re.sub(
            r"(?m)^([A-Z][A-Z0-9-]*-)\s*\n\s*([A-Z][A-Z0-9-]*)\s*$",
            r"\1\2",
            text,
        )
        code_matches = list(code_pattern.finditer(text))
        for index, match in enumerate(code_matches):
            code = match.group(1).upper()
            # Group names are uppercase too and may match ``code_pattern``.
            # Read a bounded window instead of using the preceding uppercase
            # line as a delimiter; only a nearby Cs* marker authorizes a row.
            prefix = text[max(0, match.start() - 700):match.start()]
            type_matches = list(marker_pattern.finditer(prefix))
            if not type_matches or code in excluded_codes:
                continue
            type_match = type_matches[-1]
            # The table prints the type, then its group, then the indicator code.
            group = _safe_cell(prefix[type_match.end():])
            group = re.sub(r"\b(?:TIPO|GRUPO|SIGLA|DESCRI[CÇ][AÃ]O|ESCLARECIMENTOS)\b", "", group, flags=re.IGNORECASE).strip(" -:;")
            if not group or len(group) > 180:
                continue
            next_start = code_matches[index + 1].start() if index + 1 < len(code_matches) else len(text)
            body = text[match.end():next_start]
            next_marker = marker_pattern.search(body)
            if next_marker:
                body = body[:next_marker.start()]
            lines = [
                _safe_cell(line)
                for line in body.splitlines()
                if _safe_cell(line) and not re.search(r"\b(?:Anexo|SEI)\b.*\bpg\.\s*\d+", line, re.IGNORECASE)
            ]
            if not lines:
                continue
            split_at = next((line_index for line_index, line in enumerate(lines) if explanation_starts.match(line)), None)
            description_lines = lines if split_at is None else lines[:split_at]
            clarification_lines = [] if split_at is None else lines[split_at:]
            description = _safe_cell(" ".join(description_lines))[:800]
            clarification = _safe_cell(" ".join(clarification_lines))[:6000]
            if len(description) < 8:
                continue
            canonical_key = hashlib.sha256(f"{code}|{description}".encode("utf-8")).hexdigest()[:24]
            key = (code, canonical_key)
            if key in seen:
                continue
            seen.add(key)
            definitions.append({
                "code": code,
                "indicator_type": type_match.group(1),
                "indicator_group": group,
                "official_description": description,
                "official_clarification": clarification,
                "general_guidance": "Indicador extraído do Anexo V oficial; exige revisão jurídica e contexto da competência/vínculo.",
                "source_url": source_url,
                "source_page": str(page_number),
                "canonical_key": canonical_key,
            })
    return definitions


def import_official_indicator_pdf(file_path: Path) -> ImportedCatalog:
    """Create a review-only catalog version from PT 990 Annex V downloaded locally."""
    content = file_path.read_bytes()
    if not content or len(content) > MAX_SOURCE_BYTES:
        raise ValueError("Anexo V vazio ou maior que 25 MB.")
    try:
        from pypdf import PdfReader
    except ImportError as error:  # pragma: no cover - depends on runtime
        raise RuntimeError("Importação do Anexo V requer a dependência pypdf.") from error
    try:
        reader = PdfReader(str(file_path))
        page_texts = [page.extract_text() or "" for page in reader.pages]
    except Exception as error:
        raise ValueError("Não foi possível ler o PDF oficial do Anexo V.") from error
    source_url = OFFICIAL_SOURCE_REGISTRY[0]["url"]
    definitions = _indicator_rows_from_pdf_pages(page_texts, source_url)
    if not definitions:
        raise ValueError("Nenhum indicador estruturado foi encontrado no Anexo V oficial.")
    return ImportedCatalog(
        source_name="PT 990 — Anexo V oficial do Portal IN",
        source_url=source_url,
        source_hash=hashlib.sha256(content).hexdigest(),
        definitions=definitions,
    )


def fetch_official_source(source: dict[str, str]) -> dict[str, Any]:
    """Baixa uma única URL previamente permitida e guarda cópia com hash."""
    request = Request(source["url"], headers={"User-Agent": "SOFIA-PREVI/1.0 official-source-monitor"})
    with urlopen(request, timeout=25) as response:  # nosec B310: URL vem apenas do registro acima
        content = response.read(MAX_SOURCE_BYTES + 1)
        if len(content) > MAX_SOURCE_BYTES:
            raise ValueError("Documento oficial excede o limite de 25 MB.")
        content_type = response.headers.get_content_type()
    source_hash = hashlib.sha256(content).hexdigest()
    suffix = Path(source["url"].split("?", 1)[0]).suffix or ".bin"
    target = DATA_DIR / "official_sources" / source["key"] / f"{source_hash}{suffix}"
    target.parent.mkdir(parents=True, exist_ok=True)
    if not target.exists():
        target.write_bytes(content)
    return {
        "source_key": source["key"],
        "source_url": source["url"],
        "source_hash": source_hash,
        "content_type": content_type,
        "content_length": len(content),
        "captured_at": datetime.now(timezone.utc).isoformat(),
        "local_path": str(target),
    }
